"""
地图数据管理器
"""
import json
import sqlite3
from typing import Dict, List, Optional
from pathlib import Path

try:
    import openpyxl
    from openpyxl import Workbook, load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from src.models.map import Map, Cell
from src.models.property import Property


class MapDataManager:
    """地图数据管理器"""
    
    def __init__(self):
        """初始化地图数据管理器"""
        self.supported_formats = ['json', 'xlsx', 'db']
    
    def save_map(self, map_obj: Map, format_type: str, file_path: str) -> bool:
        """
        保存地图数据
        
        Args:
            map_obj: 地图对象
            format_type: 保存格式 ('json', 'xlsx', 'db')
            file_path: 文件路径
            
        Returns:
            bool: 保存是否成功
        """
        try:
            if format_type == 'json':
                return self._save_to_json(map_obj, file_path)
            elif format_type == 'xlsx':
                return self._save_to_excel(map_obj, file_path)
            elif format_type == 'db':
                return self._save_to_database(map_obj, file_path)
            else:
                print(f"不支持的格式: {format_type}")
                return False
        except Exception as e:
            print(f"保存地图失败: {e}")
            return False
    
    def load_map(self, format_type: str, file_path: str) -> Optional[Map]:
        """
        加载地图数据
        
        Args:
            format_type: 文件格式 ('json', 'xlsx', 'db')
            file_path: 文件路径
            
        Returns:
            Optional[Map]: 地图对象，失败返回None
        """
        try:
            if format_type == 'json':
                return self._load_from_json(file_path)
            elif format_type == 'xlsx':
                return self._load_from_excel(file_path)
            elif format_type == 'db':
                return self._load_from_database(file_path)
            else:
                print(f"不支持的格式: {format_type}")
                return None
        except Exception as e:
            print(f"加载地图失败: {e}")
            return None
    
    def _save_to_json(self, map_obj: Map, file_path: str) -> bool:
        """保存为JSON格式"""
        try:
            map_data = map_obj.to_dict()
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(map_data, f, ensure_ascii=False, indent=2)
            return True
        except Exception as e:
            print(f"保存JSON失败: {e}")
            return False
    
    def _load_from_json(self, file_path: str) -> Optional[Map]:
        """从JSON格式加载"""
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                map_data = json.load(f)
            return Map.from_dict(map_data)
        except Exception as e:
            print(f"加载JSON失败: {e}")
            return None
    
    def _save_to_excel(self, map_obj: Map, file_path: str) -> bool:
        """保存为Excel格式"""
        if not OPENPYXL_AVAILABLE:
            print("openpyxl未安装，无法保存Excel格式")
            return False
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "地图数据"
            
            # 写入地图基本信息
            ws['A1'] = "地图宽度"
            ws['B1'] = map_obj.width
            ws['A2'] = "地图高度"
            ws['B2'] = map_obj.height
            ws['A3'] = "路径长度"
            ws['B3'] = map_obj.path_length
            
            # 写入路径数据
            ws['A5'] = "路径数据"
            for i, (x, y) in enumerate(map_obj.path):
                ws[f'A{6+i}'] = f"({x},{y})"
            
            # 写入格子数据
            ws['C5'] = "格子数据"
            ws['C6'] = "X"
            ws['D6'] = "Y"
            ws['E6'] = "类型"
            ws['F6'] = "路径索引"
            ws['G6'] = "路障"
            ws['H6'] = "地上金钱"
            
            row = 7
            for cell in map_obj.cells:
                ws[f'C{row}'] = cell.x
                ws[f'D{row}'] = cell.y
                ws[f'E{row}'] = cell.cell_type
                ws[f'F{row}'] = cell.path_index
                ws[f'G{row}'] = "是" if cell.roadblock else "否"
                ws[f'H{row}'] = cell.money_on_ground
                
                # 如果有房产，写入房产信息
                if cell.property:
                    ws[f'I{row}'] = f"房产等级:{cell.property.level}"
                    ws[f'J{row}'] = f"所有者:{cell.property.owner_id}" if cell.property.owner_id else "无主"
                
                row += 1
            
            wb.save(file_path)
            return True
        except Exception as e:
            print(f"保存Excel失败: {e}")
            return False
    
    def _load_from_excel(self, file_path: str) -> Optional[Map]:
        """从Excel格式加载"""
        if not OPENPYXL_AVAILABLE:
            print("openpyxl未安装，无法加载Excel格式")
            return None
        
        try:
            wb = load_workbook(file_path)
            ws = wb.active
            
            # 读取地图基本信息
            width = ws['B1'].value
            height = ws['B2'].value
            
            if not width or not height:
                print("Excel文件中缺少地图尺寸信息")
                return None
            
            # 创建地图对象
            map_obj = Map(width, height)
            
            # 读取格子数据
            row = 7
            while ws[f'C{row}'].value is not None:
                x = ws[f'C{row}'].value
                y = ws[f'D{row}'].value
                cell_type = ws[f'E{row}'].value
                path_index = ws[f'F{row}'].value
                roadblock = ws[f'G{row}'].value == "是"
                money_on_ground = ws[f'H{row}'].value or 0
                
                if x is not None and y is not None:
                    cell = map_obj.get_cell_at((x, y))
                    if cell:
                        cell.cell_type = cell_type or "empty"
                        cell.path_index = path_index or -1
                        cell.roadblock = roadblock
                        cell.money_on_ground = money_on_ground
                
                row += 1
            
            return map_obj
        except Exception as e:
            print(f"加载Excel失败: {e}")
            return None
    
    def _save_to_database(self, map_obj: Map, file_path: str) -> bool:
        """保存到数据库"""
        try:
            conn = sqlite3.connect(file_path)
            cursor = conn.cursor()
            
            # 创建地图信息表
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS map_info (
                    key TEXT PRIMARY KEY,
                    value TEXT
                )
            ''')
            
            # 插入地图基本信息
            cursor.execute('DELETE FROM map_info')
            cursor.execute('INSERT INTO map_info VALUES (?, ?)', ('width', str(map_obj.width)))
            cursor.execute('INSERT INTO map_info VALUES (?, ?)', ('height', str(map_obj.height)))
            cursor.execute('INSERT INTO map_info VALUES (?, ?)', ('path_length', str(map_obj.path_length)))
            
            # 创建路径表
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS map_path (
                    index_num INTEGER PRIMARY KEY,
                    x INTEGER,
                    y INTEGER
                )
            ''')
            
            # 插入路径数据
            cursor.execute('DELETE FROM map_path')
            for i, (x, y) in enumerate(map_obj.path):
                cursor.execute('INSERT INTO map_path VALUES (?, ?, ?)', (i, x, y))
            
            # 创建格子表
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS map_cells (
                    x INTEGER,
                    y INTEGER,
                    cell_type TEXT,
                    path_index INTEGER,
                    roadblock BOOLEAN,
                    money_on_ground INTEGER,
                    property_level INTEGER,
                    property_owner_id INTEGER,
                    PRIMARY KEY (x, y)
                )
            ''')
            
            # 插入格子数据
            cursor.execute('DELETE FROM map_cells')
            for cell in map_obj.cells:
                property_level = cell.property.level if cell.property else 0
                property_owner_id = cell.property.owner_id if cell.property else None
                
                cursor.execute('''
                    INSERT INTO map_cells VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    cell.x, cell.y, cell.cell_type, cell.path_index,
                    cell.roadblock, cell.money_on_ground,
                    property_level, property_owner_id
                ))
            
            conn.commit()
            conn.close()
            return True
        except Exception as e:
            print(f"保存数据库失败: {e}")
            return False
    
    def _load_from_database(self, file_path: str) -> Optional[Map]:
        """从数据库加载"""
        try:
            conn = sqlite3.connect(file_path)
            cursor = conn.cursor()
            
            # 读取地图基本信息
            cursor.execute('SELECT key, value FROM map_info')
            info = dict(cursor.fetchall())
            
            width = int(info.get('width', 20))
            height = int(info.get('height', 20))
            
            # 创建地图对象
            map_obj = Map(width, height)
            
            # 读取路径数据
            cursor.execute('SELECT x, y FROM map_path ORDER BY index_num')
            path_data = cursor.fetchall()
            map_obj.path = [(x, y) for x, y in path_data]
            map_obj.path_length = len(map_obj.path)
            
            # 更新格子的路径索引
            for i, (x, y) in enumerate(map_obj.path):
                cell = map_obj.get_cell_at((x, y))
                if cell:
                    cell.path_index = i
            
            # 读取格子数据
            cursor.execute('''
                SELECT x, y, cell_type, path_index, roadblock, money_on_ground,
                       property_level, property_owner_id
                FROM map_cells
            ''')
            
            for row in cursor.fetchall():
                x, y, cell_type, path_index, roadblock, money_on_ground, property_level, property_owner_id = row
                
                cell = map_obj.get_cell_at((x, y))
                if cell:
                    cell.cell_type = cell_type
                    cell.path_index = path_index
                    cell.roadblock = bool(roadblock)
                    cell.money_on_ground = money_on_ground
                    
                    # 创建房产对象
                    if property_level > 0:
                        property_obj = Property(x * map_obj.width + y, property_owner_id, property_level)
                        cell.property = property_obj
            
            conn.close()
            return map_obj
        except Exception as e:
            print(f"加载数据库失败: {e}")
            return None
    
    def get_supported_formats(self) -> List[str]:
        """
        获取支持的格式列表
        
        Returns:
            List[str]: 支持的格式列表
        """
        formats = self.supported_formats.copy()
        if not OPENPYXL_AVAILABLE:
            formats.remove('xlsx')
        return formats
    
    def validate_map(self, map_obj: Map) -> Dict[str, List[str]]:
        """
        验证地图数据
        
        Args:
            map_obj: 地图对象
            
        Returns:
            Dict[str, List[str]]: 验证结果，包含错误和警告
        """
        errors = []
        warnings = []
        
        # 检查地图尺寸
        if map_obj.width < 5 or map_obj.height < 5:
            errors.append("地图尺寸太小，建议至少5x5")
        
        if map_obj.width > 50 or map_obj.height > 50:
            warnings.append("地图尺寸较大，可能影响性能")
        
        # 检查路径
        if len(map_obj.path) < 10:
            errors.append("路径太短，建议至少10个格子")
        
        # 检查特殊格子
        bank_count = sum(1 for cell in map_obj.cells if cell.cell_type == "bank")
        shop_count = sum(1 for cell in map_obj.cells if cell.cell_type == "shop")
        jail_count = sum(1 for cell in map_obj.cells if cell.cell_type == "jail")
        
        if bank_count == 0:
            warnings.append("没有银行格子")
        if shop_count == 0:
            warnings.append("没有商店格子")
        if jail_count == 0:
            warnings.append("没有监狱格子")
        
        return {
            "errors": errors,
            "warnings": warnings
        } 